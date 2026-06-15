import os
import json
import gettext
import locale
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox as msg
from tkinter import filedialog as fd
import openpyxl
from openpyxl.styles import Font, Alignment

import inventory_manager
from AddProduct import AddProduct
from EditProduct import EditProduct
from ManageCategories import ManageCategories
from ManageSuppliers import ManageSuppliers
from StockMovement import StockMovement
from ReportsWindow import ReportsWindow

# Global translation function, defaults to identity.
_ = gettext.gettext

CONFIG_FILE = "settings.json"


# Main window of the application.
class ListProducts(tk.Tk):
    def __init__(self):
        super().__init__()

        self.db = inventory_manager.InventoryDatabase()
        self.db.create_table()

        self.var_language = tk.StringVar()

        # Load the saved language before building the UI.
        config = self.load_config()
        self.load_language(config.get("language", "en"))
        self.var_language.set("Türkçe" if config.get("language") == "tr" else "English")

        self.build_ui()
        self.update_ui()
        self.list_products()

    # Configuration helpers.
    def load_config(self):
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, encoding="utf-8") as f:
                return json.load(f)
        return {"language": "en"}

    def save_config(self, lang_code):
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump({"language": lang_code}, f)

    # Load the translation catalog for the chosen language.
    def load_language(self, lang_code):
        global _
        try:
            lang = gettext.translation(domain="messages", localedir="locales", languages=[lang_code])
            _ = lang.gettext
        except FileNotFoundError:
            _ = gettext.gettext
        try:
            if lang_code == "tr":
                locale.setlocale(locale.LC_ALL, "tr_TR.UTF-8")
            else:
                locale.setlocale(locale.LC_ALL, "en_US.UTF-8")
        except locale.Error:
            pass

    def build_ui(self):
        # Menu bar.
        self.menu = tk.Menu(self)
        self.config(menu=self.menu)
        self.file_menu = tk.Menu(self.menu, tearoff=0)
        self.actions_menu = tk.Menu(self.menu, tearoff=0)
        self.menu.add_cascade(menu=self.file_menu)
        self.menu.add_cascade(menu=self.actions_menu)

        # Top bar with the language selector.
        self.frm_top = ttk.Frame(self)
        self.frm_top.pack(padx=10, pady=10, fill="x")

        self.lbl_language = ttk.Label(self.frm_top)
        self.lbl_language.pack(side="left")
        self.combo_language = ttk.Combobox(self.frm_top, textvariable=self.var_language, state="readonly", width=12)
        self.combo_language["values"] = ["English", "Türkçe"]
        self.combo_language.pack(side="left", padx=5)
        self.combo_language.bind("<<ComboboxSelected>>", self.on_language_change)

        # Product list table.
        self.frm_list = ttk.LabelFrame(self)
        self.frm_list.pack(padx=10, pady=(0, 10), fill="both", expand=True)

        # Create the Treeview and its vertical scrollbar.
        columns = ("sku", "name", "category", "supplier", "quantity", "reorder", "price")
        self.tv = ttk.Treeview(self.frm_list, columns=columns, show="headings", height=12)
        for col in columns:
            self.tv.column(col, width=110, anchor="center")
        self.tv.pack(side="left", fill="both", expand=True, padx=5, pady=5)

        self.tv_scroll = ttk.Scrollbar(self.frm_list, orient="vertical", command=self.tv.yview)
        self.tv_scroll.pack(side="right", fill="y")
        self.tv.configure(yscrollcommand=self.tv_scroll.set)

        # Row colours for low-stock and out-of-stock products.
        self.tv.tag_configure("low", background="#fff3cd")
        self.tv.tag_configure("out", background="#f8d7da")

        # Action buttons.
        self.frm_actions = ttk.Frame(self)
        self.frm_actions.pack(padx=10, pady=(0, 10), fill="x")

        self.btn_add = ttk.Button(self.frm_actions, command=self.on_add_product)
        self.btn_add.pack(side="left", padx=2)
        self.btn_edit = ttk.Button(self.frm_actions, command=self.on_show_edit_window)
        self.btn_edit.pack(side="left", padx=2)
        self.btn_delete = ttk.Button(self.frm_actions, command=self.on_item_delete)
        self.btn_delete.pack(side="left", padx=2)
        self.btn_categories = ttk.Button(self.frm_actions, command=self.on_manage_categories)
        self.btn_categories.pack(side="left", padx=2)
        self.btn_suppliers = ttk.Button(self.frm_actions, command=self.on_manage_suppliers)
        self.btn_suppliers.pack(side="left", padx=2)
        self.btn_movement = ttk.Button(self.frm_actions, command=self.on_stock_movement)
        self.btn_movement.pack(side="left", padx=2)
        self.btn_reports = ttk.Button(self.frm_actions, command=self.on_reports)
        self.btn_reports.pack(side="left", padx=2)

        self.lbl_info = ttk.Label(self)
        self.lbl_info.pack(padx=10, pady=(0, 10), fill="x")

        # Keyboard shortcuts.
        self.tv.bind("<Double-1>", lambda e: self.on_show_edit_window())
        self.bind("<Delete>", lambda e: self.on_item_delete())
        self.bind("<Control-i>", lambda e: self.on_import_data())
        self.bind("<Control-e>", lambda e: self.on_export_data())
        self.bind("<F1>", lambda e: self.on_statistics())

    # Set every visible string through the translation function.
    def update_ui(self):
        self.title(_("Inventory & Warehouse Management"))
        self.frm_list.config(text=_("Product List"))
        self.lbl_language.config(text=_("Language"))

        self.tv.heading("sku", text=_("SKU"))
        self.tv.heading("name", text=_("Name"))
        self.tv.heading("category", text=_("Category"))
        self.tv.heading("supplier", text=_("Supplier"))
        self.tv.heading("quantity", text=_("Quantity"))
        self.tv.heading("reorder", text=_("Reorder"))
        self.tv.heading("price", text=_("Unit Price"))

        self.btn_add.config(text=_("Add Product"))
        self.btn_edit.config(text=_("Edit Product"))
        self.btn_delete.config(text=_("Delete Product"))
        self.btn_categories.config(text=_("Manage Categories"))
        self.btn_suppliers.config(text=_("Manage Suppliers"))
        self.btn_movement.config(text=_("Stock Movement"))
        self.btn_reports.config(text=_("Reports"))

        self.menu.entryconfig(1, label=_("File"))
        self.menu.entryconfig(2, label=_("Actions"))
        self.file_menu.delete(0, "end")
        self.file_menu.add_command(label=_("Import Excel"), command=self.on_import_data)
        self.file_menu.add_command(label=_("Export Excel"), command=self.on_export_data)
        self.file_menu.add_separator()
        self.file_menu.add_command(label=_("Exit"), command=self.destroy)
        self.actions_menu.delete(0, "end")
        self.actions_menu.add_command(label=_("Add Product"), command=self.on_add_product)
        self.actions_menu.add_command(label=_("Manage Categories"), command=self.on_manage_categories)
        self.actions_menu.add_command(label=_("Manage Suppliers"), command=self.on_manage_suppliers)
        self.actions_menu.add_command(label=_("Stock Movement"), command=self.on_stock_movement)
        self.actions_menu.add_command(label=_("Reports"), command=self.on_reports)

        self.lbl_info.config(text=_("Double-click to edit. Delete key to remove. F1 for statistics."))

    def on_language_change(self, event=None):
        lang_code = "tr" if self.var_language.get() == "Türkçe" else "en"
        self.load_language(lang_code)
        self.update_ui()
        self.save_config(lang_code)

    # Reload the table and colour rows by stock level (iid is the database id).
    def list_products(self):
        for item in self.tv.get_children():
            self.tv.delete(item)
        for row in self.db.get_products():
            pid, sku, name, category, supplier, quantity, reorder, price = row
            tags = ()
            if quantity == 0:
                tags = ("out",)
            elif quantity <= reorder:
                tags = ("low",)
            self.tv.insert("", "end", iid=pid, values=(sku, name, category, supplier, quantity, reorder, f"{price:.2f}"), tags=tags)

    def refresh_list(self):
        self.list_products()

    # Handlers (sub-windows are modal, list refreshes afterwards).
    def on_add_product(self):
        win = AddProduct(parent=self)
        win.grab_set()
        self.wait_window(win)
        self.refresh_list()

    def on_show_edit_window(self):
        selected = self.tv.selection()
        if not selected:
            msg.showinfo(_("Edit Product"), _("Please select a product first."), parent=self)
            return
        values = self.tv.item(selected[0], "values")
        sku, name, category, supplier, quantity, reorder, price = values
        win = EditProduct(parent=self, pid=int(selected[0]), sku=sku, name=name, category=category, supplier=supplier, quantity=quantity, reorder=reorder, price=price)
        win.grab_set()
        self.wait_window(win)
        self.refresh_list()

    def on_item_delete(self):
        selected = self.tv.selection()
        if not selected:
            msg.showinfo(_("Delete Product"), _("Please select a product first."), parent=self)
            return
        if msg.askyesno(_("Confirm"), _("Delete the selected product?"), parent=self):
            self.db.delete_product(int(selected[0]))
            self.refresh_list()

    def on_manage_categories(self):
        win = ManageCategories(parent=self)
        win.grab_set()
        self.wait_window(win)
        self.refresh_list()

    def on_manage_suppliers(self):
        win = ManageSuppliers(parent=self)
        win.grab_set()
        self.wait_window(win)
        self.refresh_list()

    def on_stock_movement(self):
        win = StockMovement(parent=self)
        win.grab_set()
        self.wait_window(win)
        self.refresh_list()

    def on_reports(self):
        win = ReportsWindow(parent=self)
        win.grab_set()
        self.wait_window(win)

    def on_statistics(self):
        count, total = self.db.get_count_and_total_value()
        message = _("Total products: {0} | Total stock value: {1:.2f}").format(count, total)
        msg.showinfo(_("Statistics"), message, parent=self)

    # Excel export.
    def on_export_data(self):
        products = self.db.get_products()
        if not products:
            msg.showerror(_("Export"), _("No data available to export."), parent=self)
            return
        path = fd.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        wb = openpyxl.Workbook()
        ws = wb.active

        # Bold header row.
        headers = ["SKU", "Name", "Category", "Supplier", "Quantity", "Reorder", "Unit Price"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # One row per product.
        for row in products:
            pid, sku, name, category, supplier, quantity, reorder, price = row
            ws.append([sku, name, category, supplier, quantity, reorder, price])
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=7):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

        # Blank spacer row, then summary formulas.
        last = ws.max_row
        ws.append([])
        ws.append(["Total Products", f"=COUNTA(A2:A{last})"])
        ws.append(["Total Stock Value", f"=SUMPRODUCT(E2:E{last},G2:G{last})"])

        widths = [12, 22, 16, 18, 10, 10, 12]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        wb.save(path)
        msg.showinfo(_("Export"), _("Export successful."), parent=self)

    # Each row is validated before it is inserted; invalid rows are skipped.
    def on_import_data(self):
        path = fd.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        categories = {c[1]: c[0] for c in self.db.get_categories()}
        suppliers = {s[1]: s[0] for s in self.db.get_suppliers()}

        imported = 0
        skipped = 0
        for values in ws.iter_rows(min_row=2, values_only=True):
            if values is None or all(v is None for v in values):
                continue
            sku, name, category, supplier, quantity, reorder, price = (list(values) + [None] * 7)[:7]
            try:
                cat_id = categories.get(category)
                sup_id = suppliers.get(supplier)
                clean = inventory_manager.ProductValidator.validate(sku, name, cat_id, sup_id, quantity, reorder, price)
                self.db.save_product(*clean)
                imported += 1
            except inventory_manager.ValidationError:
                skipped += 1
        self.refresh_list()
        message = _("Imported {0} products, skipped {1} invalid rows.").format(imported, skipped)
        msg.showinfo(_("Import"), message, parent=self)


if __name__ == "__main__":
    app = ListProducts()
    app.mainloop()
